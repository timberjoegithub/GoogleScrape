import ast
import base64
import pandas as pd
import requests
from openpyxl import Workbook, load_workbook
from env import wpAPOurl, xls,user, password,wpAPI
import datetime as dt
import json
import re
#import socket
import urllib3
from datetime import datetime

today = datetime.today().strftime('%Y-%m-%d')

#print(urllib.request.urlopen("https://www.stackoverflow.com").getcode())

def is_port_open(host, port):
#    sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
    try:
        # removehttps = re.sub(r'https://','', host)
        # removetrailing = removehttps.split('/')[0]
        isWebUp = urllib3.request("GET", host)
        #sock.close()
#        result = sock.connect_ex((host, port))
        if isWebUp.status == 200:
            return True
    except Exception as error:
        print ('Could not open port to website: ', host,  type(error))
        return False        
    
# def to_json(obj):
#      return json.dumps(obj, default=lambda obj: obj,__dict__ )

def connect_and_auth_wp( headers):
    # connect and  auth
    data_string = f"{user}:{password}"
    token = base64.b64encode(data_string.encode()).decode("utf-8")
    headers = {"Authorization": f"Basic {token}"}
    return (headers)

def check_media(filename):
    # Regex gilename to format like in WordPress media name
    #file_name_minus_extension = re.sub(r'\'|(....$)','', filename, flags=re.IGNORECASE)
    file_name_minus_extension = filename
    response = requests.get(wpAPI + "/media?search="+file_name_minus_extension, headers=headers)
    try:
        result = response.json()
        file_id = int(result[0]['id'])
        link = result[0]['guid']['rendered']
        return file_id, link
    except Exception as error:
        print('    No existing media with same name in Wordpress media folder: ' + filename)
        return False, False
    
def check_post(postname,postdate):
    # Regex gilename to format like in WordPress media name
    #file_name_minus_extension = re.sub(r'\'|(....$)','', postname, flags=re.IGNORECASE)
#    response = post_exists(postname)
#    response = requests.get(wpAPI + "/posts?search="+postname, headers=headers)
    
        # Set up query parameters (you can adjust these as needed)
    # params = {
    #     "slug": "postname" # postname,  # Specify the post name
    # }
    # Make the request
    response = requests.get(wpAPI+"/posts?search="+postname, headers=headers)
    try:
        result = response.json()
        post_id = int(result[0]['id'])
        post_date = result[0]['date']
        if postdate == post_date:
            return post_id
    except: #  Exception as error:
#        print ('Could not check to see post already exists', type(error).c)
        print('No existing post with same name: ' + postname)
    return False

def post_to_wp(title, content,  headers,date, rating,address, picslist):
    # post
    NewPost = False
    countreview = False
    addresshtml = re.sub(" ", ".",address)
    googleadress = r"<a href=https://www.google.com/maps/dir/?api=1&destination="+addresshtml + r">"+address+r"</a>" # https://www.google.com/maps/dir/?api=1&destination=760+West+Genesee+Street+Syracuse+NY+13204
    contentpics = ''
    picl = picslist[1:-1] 
    pic2 = picl.replace(",","")#re.sub(r',','',picl) #re.sub( r'[^a-zA-Z0-9]','',tempdate[1])
    pic3= pic2.replace("'","")
#    print (pic3)
    pidchop = pic3.split(" ")
    linkslist=[]
    # linkslist.clear
    print ('    Figuring out date of Post : ',title)
    if "a day" in date:
        date = datetime.timmedelta(days=-1, seconds=0, microseconds=0, milliseconds=0, minutes=0, hours=0, weeks= 0)
        newdate = dt.datetime.strptime(date_string, format).date()
    else:
        if "day" in date:
            date = datetime.timmedelta(days=-(int(re.sub( r'[^a-zA-Z]','',date))), seconds=0, microseconds=0, milliseconds=0, minutes=0, hours=0, weeks= 0)
            newdate = dt.datetime.strptime(date_string, format).date()
        else:
            if "a week" in date:
                date = datetime.timmedelta(days=0, seconds=0, microseconds=0, milliseconds=0, minutes=0, hours=0, weeks= -1)  
                newdate = dt.datetime.strptime(date_string, format).date()
            else:
                if "week" in date:
                    date = datetime.timmedelta(days=0, seconds=0, microseconds=0, milliseconds=0, minutes=0, hours=0, weeks= -(int(re.sub( r'[^a-zA-Z]','',date))))  
                    newdate = dt.datetime.strptime(date_string, format).date()
                else:
                    if "a month" in date:
                        date = datetime.timmedelta(days=0, seconds=0, microseconds=0, milliseconds=0, minutes=0, hours=0, weeks= 0, months= -1)
                        newdate = dt.datetime.strptime(date_string, format).date()
                    else:
                        if "month" in date:
                            date = datetime.timmedelta(days=0, seconds=0, microseconds=0, milliseconds=0, minutes=0, hours=0, weeks= 0, months= -(int(re.sub( r'[^a-zA-Z]','',date))))
                            newdate = dt.datetime.strptime(date_string, format).date()
                        else:
                            if "a year" in date:
                                date = datetime.timmedelta(days=0, seconds=0, microseconds=0, milliseconds=0, minutes=0, hours=0, weeks= 0, months=0, years= -1)
                                newdate = dt.datetime.strptime(date_string, format).date()
                            else:
                                if "year" in date:
                                    date = datetime.timmedelta(days=0, seconds=0, microseconds=0, milliseconds=0, minutes=0, hours=0, weeks= 0, months=0, years= -(int(re.sub( r'[^a-zA-Z]','',date))))
                                    newdate = dt.datetime.strptime(date_string, format).date()
                                else:
                                    format = '%b/%Y/%d' #specifify the format of the date_string.
                                    month = date[:3]
                                    year = date[3:]
                                    day = '01'
                                    date_string = month+'/'+year+'/'+day
                                    try: 
                                        newdate = dt.datetime.strptime(date_string, format).date()
                                        print ('    Got Date: ',str(newdate))
                                    except Exception as error:
                                        print("    An error getting date occurred:", type(error).c) # An error occurred:
                                    try:
                                        newdate = dt.datetime.strptime(date_string, format).date()
                                    except Exception as error:
                                        print("    An error getting date occurred:", type(error).c) # An error occurred:
                                    date = str(newdate)+'T22:00:00'
    try:
        post_id = check_post(title,str(date))
    except  :
    #except  Exception as error:
        print ('Could not check to see post already exists', type(error).c)
    if ( post_id == False):        
#    if ((title) != False):        
        googleadress =  r"<a href=https://www.google.com/maps/dir/?api=1&destination="+addresshtml + r">"+address+r"</a>"
        post_data = {
            "title": title,
    #        "content": address+'\n\n'+content+'\n'+rating+'\n\n' ,
            "content": googleadress+'\n\n'+content+'\n'+rating ,
            "status": "publish",  # Set to 'draft' if you want to save as a draft
            "date": date,
     #       "date": str(newdate)+'T22:00:00',
        # "author":"joesteele" 
        }
        try: 
        # response = requests.post(wpAPOurl, json=post_data, headers=headers)
        # myobj = {'json': post_data, 'headers': headers }
        # response = requests.post(wpAPOurl, json = myobj)
            headers2 = headers
            response = requests.post(wpAPOurl, json = post_data, headers=headers2)
            if ( response.status_code != 201 ):
                print ('Error: ',response)
    #            print ('json:'+ post_data+' headers :'+headers)
            # else:
            #     for loop in linkslist:
            #         print ('    Adding ', loop['link'], ' to posting')
            #         try:
            #             contentpics += r'<img src="'+ loop['link'] + r' alt="' + title +r'">' +'\n\n'
            #         except Exception as error:
            #             print("An error occurred:", type(error).__name__) # An error occurred:
            #     post_data = ()
            #     response2 = requests.post(wpAPI+"/media/" + response['id'], json = post_data, headers=headers2)
            else:
                NewPost = True
                post_id_json = response.json()
                post_id = post_id_json.get('id')
                print ('    New post is has post_id = ',post_id)
        except Exception as error:
            print("An error occurred:", type(error).__name__) # An error occurred:
        postneedsupdate = True
    else:
        print ('    Post already existed: Post ID : ',post_id)
    for pic in pidchop:
        picslice2 = pic.split("/")[-1]
        picslice = picslice2.split(".")
        picname = picslice[0] 
        caption =title
        description = title+"\n"+address
        print ('  Found Picture: ',picname)
        file_id, link = check_media(picname)
#        link = linknew['rendered']
        if (file_id) is False:
            print ('    '+picname+' was not already found in library, adding it')
            countreview = True
            image = {
                "file": open(pic, "rb"),
                "post": post_id,
                "caption": caption,
                "description": description
            }
            try:
                image_response = requests.post(wpAPI + "/media", headers=headers, files=image)
            except Exception as error:
                print("    An error uploading picture ' + picname+ ' occurred:", type(error).__name__) # An error occurred:
            if ( image_response.status_code != 201 ):
                print ('    Error- Image ',picname,' was not successfully uploaded.  response: ',image_response)
            else:
                PicDict=image_response.json()
                file_id= PicDict.get('id')
                link = PicDict.get('guid').get("rendered")
                print ('    ',picname,' was successfully uploaded to website with ID: ',file_id, link)
            try:
                linksDict = {'file_id' : file_id , 'link' : link}
                linkslist.append(linksDict)
            except Exception as error:
                print("    An error adding to dictionary " , file_id , link , " occurred:", type(error).__name__) # An error occurred:
        else:
            print ('    Photo ',picname,' was already in library and added to post with ID: ',file_id,' : ',link)
            try:
                image_response = requests.post(wpAPI + "/media/" + str(file_id), headers=headers, data={"post" : post_id})
            except Exception as error:
                print ('    Error- Image ',picname,' was not attached to post.  response: ',image_response)
            try:
                post_response = requests.post(wpAPI + "/posts/" + str(post_id), headers=headers)
                if (link in str(post_response.text)):
                    print ('    Image link for ', picname, 'already in content of post: ',post_id, post_response.text, link)
                else:
                    linkslist.append({'file_id' : file_id , 'link' : link})
                    countreview = True
            except Exception as error:
                print("    An error loading the metadata from the post " + post_response.title + ' occurred:", type(error).__name__) # An error occurred')
    #ratinghtml = post_response.text
    firstMP4 = True
    for piclink in linkslist:
        #for loop in linkslist:
        print ('    Adding ', piclink['link'], ' to posting')
        try:
            ext = (piclink['link'].split( '.')[-1] )
            if (ext == 'mp4'):
                if (firstMP4):
                    contentpics += '\n' +r'[evp_embed_video url="' + piclink['link'] + r'" autoplay="true"]'
                    firstMP4 = False
                else:
                    contentpics += '\n' +r'[evp_embed_video url="' + piclink['link'] + r'"]'           
                #[evp_embed_video url="http://example.com/wp-content/uploads/videos/vid1.mp4" autoplay="true"]
            else:
                contentpics += '\n '+r'<div class="col-xs-4"><img id="'+str(file_id)+r'"' + r'src="' + piclink['link'] + r'"></div>'
#            contentpics += '\n '+r'<img src="'+ piclink['link'] + '> \n'
            #contentpics += r'<img src="'+ piclink['link'] + r' alt="' + title +r'">' +'\n\n'
        except Exception as error:
            print("An error occurred:", type(error).__name__) # An error occurred:
    try:
        response_piclinks = requests.post(wpAPI+"/posts/"+ str(post_id), data={"content" : googleadress+'\n\n'+content+'\n'+rating  + contentpics, "featured_media" : file_id}, headers=headers)
    except Exception as error:
        print("    An error writing images to the post " + post_response.title + ' occurred:", type(error).__name__) # An error occurred')
    return NewPost

# def add_image(wpAPI,pic, picname,caption, description,headers):
#     # loop thru
#     image = {
#         "file": open(pic, "rb"),
#         "caption": caption,
#         "description": description
#     }
#     try:
#         image_response = requests.post(wpAPI + "/media", headers=headers, files=image)
#     except Exception as error:
#         print("An error occurred:", type(error).__name__) # An error occurred:
#     if ( image_response != 201 ):
#         print (image_response)
#        # print ('json:'+ post_data+' headers :'+headers)
#         print ('image response: ',image_response)
#     return image_response

# def read_from_xlsx(ws):
#     print('read from excel...')
#     return (ws)

def process_reviews(ws, headers):
    # Process
    needreversed = False
    totalcount = 2
    count = 0
    rows = list(ws.iter_rows(min_row=1, max_row=ws.max_row))
    if needreversed:
        rows = reversed(rows)
    for processrow in rows:
        if (count >= totalcount):
            print ('Exceeded the number of posts per run, exiting')
            break
#    for processrow in (ws.rows):
#        If reviewxls (comment has content) and (picsURL has connect)
        if processrow[1].value != "name":  # Skip header line of xls sheet
            print ("Processing : ",processrow[1].value)
            # ast.literal_eval(deployments) = processrow[9].value
            writtento = (ast.literal_eval(processrow[9].value))
            # Check to see if the website has already been written to according to the xls sheet, if it has not... then process
            if ((writtento["web"]) == 0) and (is_port_open(wpAPI, 443)) and (processrow[2].value != None) :
                try: 
                    NewPost = post_to_wp(processrow[1].value, processrow[2].value, headers ,processrow[7].value, processrow[3].value, processrow[8].value, processrow[5].value)
                    try:
                        writtento["web"] = 1
                        #processrow[9] = writtento[9]
#                        processrow[9].value = '"'+writtento+'"'
                        processrow[9].value = str(writtento)
                    except Exception as error:
                        print("An error occurred writing value Excel file:", type(error).__name__) # An error occurred:
                    print ('Success Posting: '+processrow[1].value)# ',processrow[1].value, processrow[2].value, headers,processrow[7].value, processrow[3].value,processrow[8].value, processrow[5].value, temp3["web"] )
                    if NewPost == True:
                        count +=1
                    try: 
                        wb.save(xls)
                    except Exception as error:
                        print("An error occurred writing Excel file:", type(error).__name__) # An error occurred:
                except: 
                    print ('Error writing post : ',processrow[1].value, processrow[2].value, headers,processrow[7].value, processrow[3].value,processrow[8].value, processrow[5].value, writtento["web"] )
    return (headers)

if __name__ == "__main__":
    countreview = False
    headers = {}
    wb = load_workbook(filename = xls)
    ws = wb['Sheet1']
    print('starting...')
   # read_from_xlsx(ws)
    print('Connect and auth to Wordpress')
    data_string = f"{user}:{password}"
    token = base64.b64encode(data_string.encode()).decode("utf-8")
    headers = {"Authorization": f"Basic {token}"}
#    print ('Headers:',headers)
    connect_and_auth_wp( headers)
    print('Processing Reviews')
#    print ('Headers:',headers)
    process_reviews(ws, headers)  
    print('Done!')