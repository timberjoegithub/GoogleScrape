"""  Script to download google maps reviews (that you create) and store them
and then on a cron'd basis post them to social media connectors """
import time
import os
from pathlib import Path
import pathlib
import re
from datetime import datetime
import ast
import base64
import datetime as dt
from urllib.request import urlretrieve
import inspect
import requests
#import json
import jsonpickle
import urllib3
import pandas as pd
from dateutil.relativedelta import relativedelta
from openpyxl import load_workbook
import instagrapi
from moviepy.editor import VideoFileClip, concatenate_videoclips
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.common.exceptions import NoSuchElementException
import tweepy
#import asyncio
#import aiohttp
import sqlalchemy
from sqlalchemy.ext.declarative import declarative_base
from sqlalchemy import null
import googlemaps
import env
#import inspect
Base = declarative_base()

##################################################################################################

class Users(Base):
    """Class representing a user config"""
    __tablename__ = 'userConfig'
    id = sqlalchemy.Column(sqlalchemy.Integer, primary_key=True)
    user = sqlalchemy.Column(sqlalchemy.String(length=11, collation="utf8"))
    instagram = sqlalchemy.Column(sqlalchemy.Boolean, default=True)
    web = sqlalchemy.Column(sqlalchemy.Boolean, default=True)
    facebook = sqlalchemy.Column(sqlalchemy.Boolean, default=True)
    tiktok = sqlalchemy.Column(sqlalchemy.Boolean, default=True)
    xtwitter = sqlalchemy.Column(sqlalchemy.Boolean, default=True)
    threads = sqlalchemy.Column(sqlalchemy.Boolean, default=True)
    yelp = sqlalchemy.Column(sqlalchemy.Boolean, default=True)
    google = sqlalchemy.Column(sqlalchemy.Boolean, default=True)
    data = sqlalchemy.Column(sqlalchemy.Boolean, default=True)
    postsperrun = sqlalchemy.Column(sqlalchemy.String(length=11, collation="utf8"))
    needreversed = sqlalchemy.Column(sqlalchemy.Boolean, default=False)
    wpAPI = sqlalchemy.Column(sqlalchemy.String(length=256, collation="utf8"))
    platform = sqlalchemy.Column(sqlalchemy.String(length=128, collation="utf8"))
    LinuxDriverLocation = sqlalchemy.Column(sqlalchemy.String(length=256, collation="utf8"))
    WindowsDriverLocation = sqlalchemy.Column(sqlalchemy.String(length=256, collation="utf8"))
    xls = sqlalchemy.Column(sqlalchemy.String(length=128, collation="utf8"))
    googleurl = sqlalchemy.Column(sqlalchemy.String(length=128, collation="utf8"))
    instagramuser = sqlalchemy.Column(sqlalchemy.String(length=32, collation="utf8"))
    instagrampass = sqlalchemy.Column(sqlalchemy.String(length=32, collation="utf8"))
    #active = sqlalchemy.Column(sqlalchemy.Boolean, default=True)

##################################################################################################

class Posts(Base):
    """Class representing the attributes of a post"""
    __tablename__ = 'posts'
    id = sqlalchemy.Column(sqlalchemy.Integer, primary_key=True)
    name = sqlalchemy.Column(sqlalchemy.String(length=256, collation="utf8"))
    comment = sqlalchemy.Column(sqlalchemy.String(length=4096, collation="utf8"))
    rating = sqlalchemy.Column(sqlalchemy.String(length=128, collation="utf8"))
    picsURL = sqlalchemy.Column(sqlalchemy.String(length=4096, collation="utf8"))
    picsLocalpath = sqlalchemy.Column(sqlalchemy.String(length=4096, collation="utf8"))
    source = sqlalchemy.Column(sqlalchemy.String(length=64, collation="utf8"))
    date = sqlalchemy.Column(sqlalchemy.String(length=64, collation="utf8"))
    address = sqlalchemy.Column(sqlalchemy.String(length=256, collation="utf8"))
    dictPostComplete = sqlalchemy.Column(sqlalchemy.String(length=128, collation="utf8"))
    #googleurl = sqlalchemy.Column(sqlalchemy.String(length=128, collation="utf8"))
    wpurl = sqlalchemy.Column(sqlalchemy.String(length=512, collation="utf8"))
    businessurl = sqlalchemy.Column(sqlalchemy.String(length=2048, collation="utf8"))
    longitude = sqlalchemy.Column(sqlalchemy.Float())
    latitude = sqlalchemy.Column(sqlalchemy.Float())
    google = sqlalchemy.Column(sqlalchemy.Boolean, default=False)
    facebook = sqlalchemy.Column(sqlalchemy.Boolean, default=False)
    instagram = sqlalchemy.Column(sqlalchemy.Boolean, default=False)
    xtwitter = sqlalchemy.Column(sqlalchemy.Boolean, default=False)
    threads = sqlalchemy.Column(sqlalchemy.Boolean, default=False)
    yelp = sqlalchemy.Column(sqlalchemy.Boolean, default=False)
    web = sqlalchemy.Column(sqlalchemy.Boolean, default=False)
    tiktok = sqlalchemy.Column(sqlalchemy.Boolean, default=False)
    #active = sqlalchemy.Column(sqlalchemy.Boolean, default=True)
    place_id = sqlalchemy.Column(sqlalchemy.String(length=126, collation="utf8"))
    googledetails = sqlalchemy.Column(sqlalchemy.String(length=4096, collation="utf8"))
    #businessurl = sqlalchemy.Column(sqlalchemy.String(length=512, collation="utf8"))
    pluscode = sqlalchemy.Column(sqlalchemy.String(length=64, collation="utf8"))
    googleurl = sqlalchemy.Column(sqlalchemy.String(length=512, collation="utf8"))
    visitdate = sqlalchemy.Column(sqlalchemy.String(length=64, collation="utf8"))


##################################################################################################

def preload():
    """
    Removes a specific file if it exists.

    Returns:
        None
    """

    file=pathlib.Path("./config/joeteststeele_uuid_and_cookie.json")
    if pathlib.Path.exists(file):
        pathlib.Path.unlink(file)
#    today = datetime.today().strftime('%Y-%m-%d')
    return

##################################################################################################

def clearlist(my_list):
    """
    Clears all elements in a list.

    Args:
        my_list (list): The list to be cleared.

    Returns:
        list: The input list with all elements cleared.
    """

    for listelement in my_list:
        listelement.clear
    return my_list

##################################################################################################

def get_auth_connect():
    """Make all connections to socials and DB, etec.."""
    connections = {}
    if env.mariadb:
        print('Connecting to MariaDB for configuration and storage')
        #from sqlalchemy import create_engine
        engine = sqlalchemy.create_engine("mysql+mysqldb://"+env.mariadbuser+":"+env.mariadbpass+
            "@"+env.mariadbserver+"/"+env.mariadbdb+"?charset=utf8mb4", echo=False)
        Session = sqlalchemy.orm.sessionmaker()
        Session.configure(bind=engine)
        session = Session()
        usersession = session.query(Users).filter(Users.user=='joesteele')
        dbuser = usersession[0]
        #for userloop in usersession:
        print("- " + dbuser.user + ' ' + dbuser.googleurl)
        connections['user'] =dbuser
        posts = session.query(Posts).all()
        connections['posts'] = posts
        connections['postssession'] = session
    if env.data:
        print('  loading XLS content data source ...')
        if os.path.exists(env.xls):
            wb = load_workbook(filename = env.xls)
            xls_wb_df = pd.read_excel(env.xls)
        elif os.path.exists('./GoogleScrape/'+ env.xls):
            wb = load_workbook(filename = './GoogleScrape/'+ env.xls)
            xls_wb_df = pd.read_excel('./GoogleScrape/'+ env.xls)
        else:
            input("Not able to find xls file Press any key to continue...")
        ws = wb['Sheet1']
        #xls_wb_df = pd.read_excel(xls)
        connections |= {'xlsdf':xls_wb_df,'data':ws,'datawb':wb}
#        connections.update({'data':ws})
#        connections.update({'datawb':wb})

        sheet = wb.active
        c_row = sheet.max_row
        c_column = sheet.max_column
        #out_data=[]
        #import inspect
        local_dict=Posts
        column_list = []
        my_list = []
        my_list_data = []
        my_post = Posts
        my_dict = {}
        for x in inspect.getmembers(Posts):
            if not (x[0].startswith('_') or 'metadata' in x[0]):
                column_list.append(x[0])
        # solumn_list2[0][0]
        #solumn_list2 = tuple(x for x in solumn_list2 if ('metadata' not in solumn_list2[0][0]) or (not solumn_list2[0][0].startswith('_')))
#        for a in range(1, c_row + 1):
        for a in range(1, c_row + 1):
            for b in range(2, c_column+1):
                ob = sheet.cell(row=a, column=b)
                my_list.append(ob.value)
#                my_post.getattr(Posts, column_list[b]) == ob.value #my_post.(eval(b,column_list))
                #my_dict[b] = ob.value
            my_list_data.append(str(my_list))
            print (my_list)
            my_list = []
            #out_data.append(str(my_dict))
        connections['xls_list'] = my_list_data
#        connections['xls_list'] = out_data
    if env.instagram:
        print('  Connecting to Instagram ...')
        instasessionclient = instagrapi.Client()
        instasessionclient.login(env.instagramuser, env.instagrampass)
        connections['instagram'] = instasessionclient
    if env.facebook:
        print('  Connecting to facebook ...')
        # page_id_1 = facebookpageID
        # facebook_access_token = 'paste-your-page-access-token-here'
        # image_url = 'https://graph.facebook.com/{}/photos'.formatting(page_id_1)
        # image_location = 'http://image.careers-portal.co.za/f_output.jpg'
        # img_payload = {
        # 'message': msg,
        # 'url': image_location,
        # 'access_token': facebook_access_token
        # }
        # #Send the POST request
        # r = requests.post(image_url, data=img_payload)
        # print(r.text)
        connections['facebook'] = posts
    if env.yelp:
        print('  Connecting to yelp ...')
    if env.xtwitter:
        print('  Connecting to xtwitter ...')
    if env.threads:
        print('  Connecting to threads ...')
        threadssessionclient = instagrapi.Client()
        threadssessionclient.login(env.instagramuser, env.instagrampass)
        connections |= {'threads':threadssessionclient}
    if env.web:
        print('  Connecting to joeeatswhat.com ...')
        data_string = f"{env.user}:{env.password}"
        token = base64.b64encode(data_string.encode()).decode("utf-8")
        connections['web'] = {"Authorization": f"Basic {token}"}
    else:
        connections['web'] = ""
    if env.tiktok :
        print('  Connecting to Instagram ...')
    return connections

##################################################################################################

def get_twitter_conn_v1(api_key, api_secret, access_token, access_token_secret) -> tweepy.API:
    """Get twitter conn 1.1"""
    auth = tweepy.OAuth1UserHandler(api_key, api_secret)
    auth.set_access_token(
        access_token,
        access_token_secret,
    )
    return tweepy.API(auth)

##################################################################################################

def get_twitter_conn_v2(api_key, api_secret, access_token, access_token_secret) -> tweepy.Client:
    """Get twitter conn 2.0"""
    return tweepy.Client(
        consumer_key=api_key,
        consumer_secret=api_secret,
        access_token=access_token,
        access_token_secret=access_token_secret,
    )

##################################################################################################

def get_hastags(address, name, hashtype):
    """
    Generates hashtags based on the address, name, and type provided.

    Args:
        address (str): The address related to the content.
        name (str): The name associated with the content.
        hashtype (str): The type of hashtags to generate.

    Returns:
        str: The generated hashtags based on the input parameters.
    """

    name_no_spaces = re.sub( r'[^a-zA-Z]','',name)
    addressdict = address.rsplit(r' ',3)
    zip_code = addressdict[3]
    state = addressdict[2]
    city =  re.sub( r'[^a-zA-Z]','',addressdict[1])
    if 'short' in hashtype:
        defaulttags = '#'+name_no_spaces+' #foodie #food #joeeatswhat @timberjoe'
    else:
        defaulttags = "\n\n\n#"+name_no_spaces+\
            " #foodie #music #food #travel #drinks #instagood #feedme #joeeatswhat @timberjoe"
    citytag = "#"+city
    statetag = "#"+state
    ziptag = "#"+zip_code
    if statetag == 'FL':
        statetag += '#Florida'
    if statetag == 'OR':
        statetag += '#Oregon'
    if statetag == 'MA':
        statetag += '#Massachusetts'
    return defaulttags+" "+citytag+" "+statetag+" "+ziptag+" "

##################################################################################################

# Grab a count of how far we need to scroll
def counter_google(driver):
    """
    Counts the number of Google search results pages.

    Args:
        driver: The Selenium WebDriver instance.

    Returns:
        int: The total number of search result pages.
    """

    result = driver.find_element(By.CLASS_NAME,'Qha3nb').text
    result = result.replace(',', '')
    result = result.split(' ')
    result = result[0].split('\n')
    return int(result[0]) // 10 + 1

##################################################################################################

def make_montage_video_from_google(inphotos):
    """
    Creates a montage video from a list of input photos.

    Args:
        inphotos (list): List of paths to input photo files.

    Returns:
        tuple: A tuple containing the path to the output video file and a boolean indicating succes
    """

# Load the photos from the folder
# Set the duration of each photo to 2 seconds
    if not inphotos:
        return False, False
    directory = inphotos[0].rsplit(r'/', 1)
    folder = directory[0]
    output = folder+"/montage.mp4"
    if not os.path.exists(output) and len(inphotos) >1:
        video = VideoFileClip(inphotos[0])
        for photo in inphotos :
            #clip = VideoFileClip("myHolidays.mp4").subclip(50,60)
            clip = VideoFileClip(photo)
            # Concatenate the photos into a clip
            if ".jpg" in photo:
                clip.duration = 2
            try:
                video = concatenate_videoclips([video, clip], method="compose")
            except AttributeError as error:
                print("  An error occurred :", type(error).__name__) # An error occurred:
        # Load the audio file
        #    audio = AudioFileClip("audio.mp3")
        # Set the audio as the soundtrack of the montage
        #    montage = montage.set_audio(audio)
        # Write the final montage to a file
        outputvideo = video.write_videofile(folder+"/montage.mp4", fps=24)
    else:
        outputvideo = False
    return outputvideo, output

##################################################################################################

def is_docker():
    """
    Checks if the code is running in a Docker container.

    Returns:
        bool: True if running in a Docker container, False otherwise.
    """
    cgroup = Path('/proc/self/cgroup')
    #print (cgroup.read_text())
    return Path('/.dockerenv').is_file() or cgroup.is_file() and 'docker' in \
            cgroup.read_text(encoding="utf-8")

##################################################################################################

def post_facebook_video(group_id, video_path, auth_token, title, content, date, rating, address):
    """
    Posts a video to a Facebook group with specified details.

    Args:
        group_id (str): The ID of the Facebook group.
        video_path (list): List of paths to the video files to be uploaded.
        auth_token (str): The authentication token for posting to Facebook.
        title (str): The title of the video.
        content (str): Additional content to be included in the post.
        date (str): The date of the post.
        rating (str): The rating associated with the post.
        address (str): The address related to the post.

    Returns:
        dict or bool: The response JSON if successful, False if an error occurs.
    """

    url = f"https://graph-video.facebook.com/{group_id}/videos?access_token=" + auth_token
    addresshtml = re.sub(" ", ".",address)
    files = {eachfile: open(eachfile, 'rb') for eachfile in video_path}
    data = { "title":title,"description" : title + "\n"+ address+"\nGoogle map to destination: "
            r"https://www.google.com/maps/dir/?api=1&destination="+addresshtml +"\n\n"+ content +
            "\n"+rating+"\n"+date+"\n\n"+ get_hastags(address, title, 'long')+
            "\n\nhttps://www.joeeatswhat.com"+"\n\n","published" : True,
            "alt_text" : title
    }
    try:
        r = requests.post(url, files=files, data=data,timeout=env.request_timeout).json()
    except AttributeError as error:
        print("    An error getting date occurred:", error) # An error occurred:
        r = False
    time.sleep(env.facebooksleep)
    return r

##################################################################################################

def get_google_data(driver, local_outputs):
    """
    Retrieves data from Google Maps including name, address, and review details.

    Args:
        driver: The Selenium WebDriver instance.
        local_outputs: Output information.

    Returns:
        list: A list of data extracted from Google Maps.
    """
    print('get google data...')
    # Click on more botton on each text reviews
    more_elemets = driver.find_elements(By.CSS_SELECTOR, '.w8nwRe.kyuRq')
    for list_more_element in more_elemets:
        list_more_element.click()
    # Find Pictures that have the expansion indicator to see the rest of the pictures under
    #    them and click it to expose them all
    more_pics = driver.find_elements(By.CLASS_NAME, 'Tya61d')
    for list_more_pics in more_pics:
        if 'showMorePhotos' in  list_more_pics.get_attribute("jsaction") :
            print('    Found extra pics')
            list_more_pics.click()
    elements = driver.find_elements(By.CLASS_NAME, 'jftiEf')
    lst_data = []
    for data in elements:
        name = data.find_element(By.CSS_SELECTOR, 'div.d4r55.YJxk2d').text
        try:
            address = data.find_element(By.CSS_SELECTOR, 'div.RfnDt.xJVozb').text
        except NoSuchElementException :
            address = 'Unknonwn'
        print ('Name of location: ',name, '   Address:',address)
        try:
            visitdate = data.find_element(By.CSS_SELECTOR, 'span.rsqaWe').text
        except NoSuchElementException :
            visitdate = "Unknown"
        print('  Visited: ',visitdate)
        try:
            text = data.find_element(By.CSS_SELECTOR, 'div.MyEned').text
        except NoSuchElementException :
            text = ''
        try:
            score = data.find_element(By.CSS_SELECTOR, 'span.kvMYJc').get_attribute("aria-label")
        #find_element(By.CSS_SELECTOR,'aria-label').text #)  ##QA0Szd > div > div > div.w6VYqd >
        #  div:nth-child(2) > div > div.e07Vkf.kA9KIf > div > div > div.m6QErb.DxyBCb.kA9KIf.dS8AEf
        #  > div.m6QErb > div:nth-child(3) > div:nth-child(2) > div > div:nth-child(4) > div.DU9Pgb
        #  > span.kvMYJc
        except NoSuchElementException  as error:
            score = "Unknown"
            print ('Error: ',error)
        more_specific_pics = data.find_elements(By.CLASS_NAME, 'Tya61d')
    #  Grab more info from google maps entry on this particular review
        if len(local_outputs['postssession'].query(Posts).filter(Posts.name == name,Posts.google\
                is not True).all()) == 0 or env.forcegoogleupdate or env.block_google_maps is not\
                True:
            gmaps = googlemaps.Client(env.googleapipass)
            place_ids = gmaps.find_place(name+address, input_type = 'textquery', fields='')
            if len(place_ids['candidates']) == 1 :
                place_id = place_ids['candidates'][0]['place_id']
                details = gmaps.place(place_id)
            # Get place details
                try:
                    businessurl = details['result']['website']
                    latitude = details['result']['geometry']['location']['lat']
                    longitude = details['result']['geometry']['location']['lng']
                    pluscode = details['result']['plus_code']['compound_code']
                    googleurl = details['result']['url']
                    database_update_row(name,"businessurl",businessurl,"onlyempty",local_outputs)
                    database_update_row(name,"latitude",latitude,"onlyempty",local_outputs)
                    database_update_row(name,"longitude",longitude,"onlyempty",local_outputs)
                    database_update_row(name,"pluscode",pluscode,"onlyempty",local_outputs)
                    database_update_row(name,"googleurl",googleurl,"onlyempty",local_outputs)
                    database_update_row(name,"place_id",place_id,"onlyempty",local_outputs)
                    database_update_row(name,"googledetails",details,"onlyempty",local_outputs)
                    database_update_row(name,"google",True,"onlyempty",local_outputs)
                except KeyError as error:
                    print('Error writing business details from google maps : ',error)
        else:
            print ('  Post was already in database, skipping update unless you activate override')
        database_update_row(name,"google",True,"forceall",local_outputs)
        pics= []
        pics2 = []
        # check to see if folder for pictures and videos already exists, if not, create it
        cleanname = re.sub( r'[^a-zA-Z0-9]','', name)
        if not os.path.exists('./Output/Pics/'+cleanname):
            os.makedirs('./Output/Pics/'+cleanname)
        # Walk through all the pictures and videos for a given review
        for lmpics in more_specific_pics:
            # Grab URL from style definiton (long multivalue string), and remove the -p-k so that
            #   it is full size
            urlmedia = re.sub(r'=\S*-p-k-no', '=-no', (re.findall(r"['\"](.*?)['\"]",
                lmpics.get_attribute("style")))[0])
            print ('    Pic URL : ',urlmedia)
            pics.append(urlmedia)
            # Grab the name of the file and remove all spaces and special charecters to name the
            #    folder
            filename = re.sub( r'[^a-zA-Z0-9]','', str(lmpics.get_attribute("aria-label")))
            if lmpics == more_specific_pics[0]:
                lmpics.click()
                time.sleep(2)
                #iframe = driver.find_element(By.TAG_NAME, "iframe")
                tempdate = str((driver.find_element(By.CLASS_NAME,'mqX5ad')).text).rsplit("-",1)
                visitdate = re.sub( r'[^a-zA-Z0-9]','',tempdate[1])
                #print ('  Visited: ',visitdate)
            # Check to see if it has a sub div, which represents the label with the video length
            # displayed, this will be done
            # because videos are represented by pictures in the main dialogue, so we need to click
            # through and grab the video URL
            if lmpics.find_elements(By.CSS_SELECTOR,'div.fontLabelMedium.e5A3N') :
                ext='.mp4'
                lmpics.click()
                time.sleep(2)
                # After we click the right side is rendered in an inframe, Store iframe web element
                iframe = driver.find_element(By.TAG_NAME, "iframe")
                # switch to selected iframe
                driver.switch_to.frame(iframe)
                # Now find button and click on button
                video_elements = driver.find_elements(By.XPATH ,'//video') #.get_attribute('src')
                urlmedia = str((video_elements[0]).get_attribute("src"))
                # return back away from iframe
                driver.switch_to.default_content()
            else:
                # The default path if it is not a video link
                ext='.jpg'
            # Add the correct extension to the file name
            filename = filename+ext
            # Test to see if file already exists, and if it does not grab the media and store it
            #   in location folder
            if not os.path.exists('./Output/Pics/'+cleanname+'/'+visitdate):
                os.makedirs('./Output/Pics/'+cleanname+'/'+visitdate)
            if not os.path.isfile('./Output/Pics/'+cleanname+'/'+visitdate+'/'+filename):
                urlretrieve(urlmedia, './Output/Pics/'+cleanname+'/'+visitdate+'/'+filename)
            # Store the local path to be used in the excel document
            pics_local_path = "./Output/Pics/"+cleanname+"/"+visitdate+'/'+filename
            pics2.append(pics_local_path)
        if pics2:
            make_montage_video_from_google(pics2)
            pics2.append("./Output/Pics/"+cleanname+"/"+visitdate+'/'+'montage.mp4')
        dict_post_complete= {'google':1,'web':0,'yelp':0,'facebook':0,'xtwitter':0,
            'instagram':0,'tiktok':0}
        lst_data.append([name , text, score,pics,pics2,"GoogleMaps",visitdate,address,
            dict_post_complete])
        database_update_row(name,"visitdate",visitdate,"forceall",local_outputs)
    return lst_data

##################################################################################################

# Do the google_scroll
def google_scroll(counter_google_scroll,driver):
    """
    Scrolls down a Google search results page a specified number of times.

    Args:
        counter_google_scroll (int): The number of times to scroll down the page.
        driver: The Selenium WebDriver instance.

    Returns:
        int: The result of the last scroll operation.

    Raises:
        AttributeError: If an error occurs during scrolling.
    """
    print('google_scroll...', end ="")
    time.sleep(3)
    scrollable_div = driver.find_element(By.XPATH,
        '//*[@id="QA0Szd"]/div/div/div[1]/div[2]/div/div[1]/div/div/div[5]/div[2]')
#        '//*[@id="QA0Szd"]/div/div/div[1]/div[2]/div/div[1]/div/div/div[2]/div[10]/div')
    for _i in range(counter_google_scroll):
        try:
            google_scroller = driver.execute_script(
                'document.getElementsByClassName("dS8AEf")[0].\
                    scrollTop=document.getElementsByClassName("dS8AEf")[0].scrollHeight',
                    scrollable_div
            )
            time.sleep(3)
            print('.')
        except AttributeError  as e:
            print(f"Error while google_scroll: {e}")
            break
    print ('')
    return google_scroller

##################################################################################################

def write_to_xlsx2(data, local_outputs):
    """
    Writes data to an Excel file and updates the database with new entries.

    Args:
        data: Data to be written to the Excel file.
        local_outputs: Output information.

    Returns:
        Data: The data that was written to the Excel file.
    """

    print('write to excel...')
    sqlalchemy.null()
    cols = ["name", "comment", 'rating','picsURL','picsLocalpath','source','date','address',
        'dictPostComplete']
    cols2 = ["num","name", "comment", 'rating','picsURL','picsLocalpath','source','date',
        'address','dictPostComplete']
    df = pd.DataFrame(data, columns=cols)
    df2 = pd.DataFrame(local_outputs['xlsdf'].values, columns=cols2)
    #df2 = df1.where((pd.notnull(df)), None)  # take out NAN problems
    #df3.astype(object).where(pd.notnull(df2), None)
    print ('Dropped items not included in sync to database: ',df2.dropna(inplace=True))
    rows = list(data)
    if env.needreversed:
        rows = reversed(rows)
    #jsonposts = json.dumps(local_outputs['posts'], default=Posts)
    print("Encode Object into JSON formatted Data using jsonpickle")
    jsonposts = jsonpickle.encode(local_outputs['posts'], unpicklable=False)
    for processrow in df2.values:
        if processrow[1] in df.values:
            print ('  Row ',processrow.id,' ', processrow.name ,'  already in XLS sheet')
            d2_row = Posts(name=processrow.name ,comment=processrow.comment,rating=\
                processrow.rating,picsURL=processrow.picsURL,pics_local_path=processrow.\
                picsLocalpath,source=processrow.source,date=processrow.date,address=processrow.\
                address,dict_post_complete=processrow.dictPostComplete)
        elif processrow.name is not None:
# Create a Python dictionary object with all the column values
# d_row = {'name':processrow.name ,'comment':processrow.comment,'rating':processrow.rating,
#     'picsURL':processrow.picsURL,'pics_local_path':processrow.picsLocalpath, 'source':\
#      processrow.source,'date':processrow.date,'address':processrow.address,'dict_post_complete'\
#      :processrow.dictPostComplete}
            d2_row = Posts(name=processrow.name ,comment=processrow.comment,rating=processrow.\
                rating,picsURL=processrow.picsURL,pics_local_path=processrow.picsLocalpath,\
                source=processrow.source,date=processrow.date,address=processrow.address,\
                dict_post_complete=processrow.dictPostComplete)
            print ('  Row ',processrow[0],' ', processrow.name ,'  added to XLS sheet')
        # Append the above Python dictionary object as a row to the existing pandas DataFrame
        # Using the DataFrame.append() function
        try:
            if processrow.name in jsonposts : #local_outputs['posts']):
                print ('  Row ',processrow[0],' ', processrow.name ,'  already in Database')
            else:
                local_outputs['postssession'].add(d2_row)
                local_outputs['postssession'].commit()
                print ('  Row ',processrow[0],' ', processrow.name ,'  added to Database')
        except AttributeError as error:
            print('    Not able to write to post data table: ' , type(error))
            local_outputs['postssession'].rollback()
            raise
    df.to_excel(env.xls)
    return data

##################################################################################################

def write_to_database(data, local_outputs):
    """
    Writes data to the database and updates the database with new entries.

    Args:
        data: Data to be written to the database.
        local_outputs: Local output information.

    Returns:
        Data: The data that was written to the database.
    """
    column_list=[]
    print('write to database ...')
    for x in inspect.getmembers(Posts):
        if not (x[0].startswith('_') or 'metadata' in x[0]):
            column_list.append(x[0])
    cols = ["name", "comment", 'rating','picsURL','picsLocalpath','source','date','address',
        'dictPostComplete','visitdate']
    cols2 = ["num","name", "comment", 'rating','picsURL','picsLocalpath','source','date',
        'address','dictPostComplete','visitdate']
    #df = pd.DataFrame(local_outputs["xls"], columns=cols)
#    df = pd.DataFrame(local_outputs['xlsdf'])
    df = pd.DataFrame(local_outputs['xlsdf'].values, columns=column_list)
    df2 = pd.DataFrame(local_outputs['posts'])
    # print ('Dropped items not included in sync to database: ',df2.dropna(inplace=True))
#    rows = list(data)
    # if env.needreversed:
    #     rows = reversed(rows)
    #jsonposts = json.dumps(local_outputs['posts'], default=Posts)
    #print("Encode Object into JSON formatted Data using jsonpickle")
    jsonposts = jsonpickle.encode(local_outputs['posts'], unpicklable=False)
    for processrow in data:
        if processrow.name in df.values:
            print ('  Row ',processrow.id,' ', processrow.name ,'  already in database')
            d2_row = Posts(name=processrow.name ,comment=processrow.comment,rating=processrow.\
                rating,picsURL=processrow.picsURL,pics_local_path=processrow.picsLocalpath,\
                source=processrow.source,date=processrow.date,address=processrow.address,\
                dict_post_complete=processrow.dictPostComplete)
        elif processrow.name is not None:
# Create a Python dictionary object with all the column values
# d_row = {'name':processrow.name ,'comment':processrow.comment,'rating':processrow.rating,
#     'picsURL':processrow.picsURL,'pics_local_path':processrow.picsLocalpath, 'source':
#      processrow.source,'date':processrow.date,'address':processrow.address,'dict_post_complete'
#      :processrow.dictPostComplete}
            d2_row = Posts(name=processrow.name ,comment=processrow.comment,rating=processrow.\
                rating,picsURL=processrow.picsURL,pics_local_path=processrow.picsLocalpath,\
                source=processrow.source,date=processrow.date,address=processrow.address,\
                dict_post_complete=processrow.dictPostComplete)
            print ('  Row ',processrow[0],' ', processrow.name ,'  added to XLS sheet')
        # Append the above Python dictionary object as a row to the existing pandas DataFrame
        # Using the DataFrame.append() function
        try:
            if processrow.name in jsonposts : #local_outputs['posts']):
                print ('  Row ',processrow[0],' ', processrow.name ,'  already in Database')
            else:
                local_outputs['postssession'].add(d2_row)
                local_outputs['postssession'].commit()
                print ('  Row ',processrow[0],' ', processrow.name ,'  added to Database')
        except AttributeError as error:
            print('    Not able to write to post data table: ' , type(error))
            local_outputs['postssession'].rollback()
            raise
    df.to_excel(env.xls)
    return data

##################################################################################################

def database_update_row(review_name, column_name, column_value, update_style, local_outputs):
    """
    Updates a row in the database based on the specified review name, column name, value, 
    and update style.

    Args:
        review_name (str): The name of the review to update.
        column_name (str): The name of the column to update.
        column_value: The value to update the column with.
        update_style (str): The style of update to perform.
        local_outputs: Local output information.

    Returns:
        bool: True if the row was successfully updated, False otherwise.
    """

    try:
        if update_style == "forceall" and column_value is not False:
            local_outputs['postssession'].query(Posts).filter(Posts.name == review_name).update\
                        ({column_name : column_value})
            print ('    Force Updated ',column_name, ' to: ',column_value)
        elif update_style == "onlyempty"  and column_value is not False:
            postval = local_outputs['postssession'].query(Posts).filter(Posts.name == review_name,\
                            getattr(Posts,column_name).is_not(null())).all()
            if len(postval) == 0 :
                local_outputs['postssession'].query(Posts).filter(Posts.name == review_name).update\
                    ({column_name : column_value})
                print ('    Updated blank ',postval ,' on value',column_name, ' to: ',column_value)
        elif update_style == "toggletrue":
            postval = local_outputs['postssession'].query(Posts).filter(Posts.name == review_name,\
                            getattr(Posts,column_name).is_not(1)).all()
            local_outputs['postssession'].query(Posts).filter(Posts.name == review_name).update\
                    ({column_name : "1"})
            print ('    Updated ',column_name, ' on value: ',postval[0].column_value, ' to: ',\
                column_value)
    except AttributeError as error:
        print("    Not able to write to post data table to update ",review_name," ",column_name,"\
            to: ",column_value , type(error), error)
        local_outputs['postssession'].rollback()
        raise
    else:
        local_outputs['postssession'].commit()
    return True

##################################################################################################

def check_wordpress_media(filename, headers):
    """
    Checks if a media file exists in the WordPress media folder.

    Args:
        filename (str): The name of the media file to check.
        headers: Additional headers for the request.

    Returns:
        tuple: A tuple containing the ID and link of the media file if found, otherwise 
        (False, False).
    """

    file_name_minus_extension = filename
    response = requests.get(env.wpAPI + "/media?search="+file_name_minus_extension,\
                        headers=headers,timeout=env.request_timeout)
    try:
        result = response.json()
        if result:
            file_id = int(result[0]['id'])
            link = result[0]['guid']['rendered']
            return file_id, link
        else:
            return False,False
    except AttributeError:
        print('    No existing media with same name in Wordpress media folder: '+filename)
        return (False, False)

##################################################################################################

def check_is_port_open(host, port):
    """
    Checks if a port on a host is open.

    Args:
        host (str): The host to check for an open port.
        port (int): The port number to check.

    Returns:
        bool: True if the port is open, False otherwise.
    """

    try:
        is_web_up = urllib3.request("GET", host, timeout=env.request_timeout)
        if is_web_up.status == 200:
            return True
    except AttributeError as error:
        print ('Could not open port to website: ', host,  type(error))
        return False

##################################################################################################

def get_wordpress_post_id_and_link(postname, headers2):
    """
    Retrieves the ID and link of a WordPress post based on the post name.

    Args:
        postname (str): The name of the WordPress post to retrieve.
        headers2: Additional headers for the request.

    Returns:
        tuple: A tuple containing the ID and link of the WordPress post.
    """

    response = requests.get(env.wpAPI+"/posts?search="+postname, headers=headers2,timeout=env.request_timeout)
    result = response.json()
    if len(result) > 0:
        return int(result[0]['id']), result[0]['link']
    print('No existing post with same name: ' + postname)
    return False, False

##################################################################################################

def check_wordpress_post(postname, postdate, headers2):
    """
    Checks if a WordPress post with the given name and date exists.

    Args:
        postname (str): The name of the WordPress post to check.
        postdate: The date of the WordPress post to check.
        headers2: Additional headers for the request.

    Returns:
        tuple: A tuple containing the ID and link of the existing post if found, 
        otherwise (False, False).
    """

    response = requests.get(env.wpAPI+"/posts?search="+postname, headers=headers2,timeout=env.request_timeout)
    result = response.json()
    if len(result) > 0 and postdate == result[0]['date']:
        return int(result[0]['id']), result[0]['link']
    print('No existing post with same name: ' + postname)
    return False, False

##################################################################################################

# Function to get the featured photo ID of a WordPress post
def get_wordpress_featured_photo_id(post_id):
    """ Get photoid from featured photo

    Args:
        post_id (_type_): _description_

    Returns:
        _type_: _description_
    """
    # Make a GET request to the WordPress REST API to retrieve media details
    response = requests.get(f"{env.wpAPOurl}/{post_id}",timeout=env.request_timeout)
    # Check if the request was successful
    if response.status_code == 200:
        # Parse the JSON response
        media_items = response.json()

        # Loop through the media items associated with the post
           
        # for item in media_items:
        #     # Check if the media item is the featured image
        #     if item.get('post', None) == int(post_id):
        #         # Return the ID of the featured image
        return  media_items['featured_media']
    # If the request failed or the featured image was not found, return None
    return None
# # Example usage
# featured_photo_id = get_featured_photo_id(POST_ID)
# print(f"The featured photo ID for post {POST_ID} is: {featured_photo_id}")

##################################################################################################

def post_to_x2(title, content, headers,date, rating, address, picslist,local_outputs):
    """
    Post to x2.

    This function posts content to a social media platform using the provided data.

    Args:
        title (str): The title of the post.
        content (str): The content of the post.
        date (str): The date of the post.
        rating (int): The rating of the post.
        address (str): The address associated with the post.
        picslist (list): A list of pictures for the post.
        instasession: The Instagram session for posting.

    Returns:
        str: The media ID of the posted content.
    """
    pics = ((picslist[1:-1]).replace("'","")).split(",")
    # Replace the following strings with your own keys and secrets
    CONSUMER_KEY = env.x_consumer_key
    CONSUMER_SECRET = env.x_consumer_secret
    ACCESS_TOKEN = env.x_access_token
    ACCESS_TOKEN_SECRET = env.x_access_token_secret
    # Authenticate to Twitter
    auth = tweepy.OAuthHandler(CONSUMER_KEY, CONSUMER_SECRET)
    auth.set_access_token(ACCESS_TOKEN, ACCESS_TOKEN_SECRET)
    # Create an API object to use the Twitter API
    #api = tweepy.API(auth)
    img_list = pics
    imgs_vid = []
    imgs_pic = []
    client_v1 = get_twitter_conn_v1(CONSUMER_KEY,CONSUMER_SECRET,ACCESS_TOKEN,ACCESS_TOKEN_SECRET)
    client_v2 = get_twitter_conn_v2(CONSUMER_KEY,CONSUMER_SECRET,ACCESS_TOKEN,ACCESS_TOKEN_SECRET)
    for img in img_list:
        if 'montage.mp4' in img:
            imgs_vid.append(img.strip())
        else:
            imgs_pic.append(img.strip())
    if imgs_vid:
        try:
            video_path = imgs_vid[0]
            # Message to post along with the video
            attrib_list = local_outputs['postssession'].query(Posts).filter(Posts.name == title)\
                    .all()
            business_url = attrib_list[0].businessurl
            wpurl = attrib_list[0].wpurl
            if wpurl:  # Don't post of website URL does not exist yet
                if business_url:  #Sometimes Business URL does not exist, so account for it
                    status_message=str(title)+': My Review - '+wpurl+'\n Business website: '+\
                        business_url + '\n'
                else:
                    status_message = str(title) + ': My Review - '+ wpurl +  '\n'
                status_message2  = status_message +' '+str(get_hastags(address, title, 'short'))+' '
                status_message_short = status_message2[:279]
                # Upload video
                media = client_v1.media_upload(filename=video_path)
                # Post tweet with video
                # tweetlat = (local_outputs['postssession'].query(Posts).filter(Posts.name == \
                #       title).all())[0].latitude
                # tweetlong = (local_outputs['postssession'].query(Posts).filter(Posts.name == \
                #       title).all())[0].longitude
                if media.processing_info['state'] != 'failed':
                    client_v2.create_tweet(text=status_message_short,media_ids=[media.media_id])
#                    client_v2.create_tweet(text=status_message_short, lat=tweetlat , \
#                           long=tweetlong ,media_ids=[media.media_id])
                else:
                    print ('Problem uploading video to twitter: ',media.processing_info['error'])
                    return False
            else:
                print ('  Skipping xtwitter post because wordpress URL is missing')
        except AttributeError  as error:
            print("AttributeError     An error occurred:",error) # An error occurred:
    else:
        print ('    xtwitter: No video available, skipping twitter post for : ',title)
        return False
    time.sleep(env.facebooksleep)
    return True

##################################################################################################

def post_facebook3(title, content,headers, date, rating, address, picslist, local_outputs):
    """
    Post to Facebook3.

    This function posts content to Facebook using the provided data.

    Args:
        title (str): The title of the post.
        content (str): The content of the post.
        date (str): The date of the post.
        rating (int): The rating of the post.
        address (str): The address associated with the post.
        picslist (list): A list of pictures for the post.
        instasession: The Instagram session for posting.

    Returns:
        bool: Indicates if the post was successfully made.
    """
    pics = ((picslist[1:-1]).replace("'","")).split(",")
    group_id = env.facebookpageID
    auth_token = env.facebookpass
    imgs_id = []
    imgs_vid = []
    imgs_pic = []
    img_list = pics
    attrib_list = local_outputs['postssession'].query(Posts).filter(Posts.name == title).all()
    business_url = attrib_list[0].businessurl
    wpurl = attrib_list[0].wpurl
    if wpurl and (attrib_list[0].picsLocalpath != '[]'):
        if business_url:
            status_message = (
                    f'{str(title)}: My Review - {wpurl} \n Business website: {business_url}'\
                        + ' \n\n') + content
        else:
            status_message = f'{str(title)}: My Review - ' + wpurl + ' \n\n' + content
        for img in img_list:
            if 'montage.mp4' in img:
                imgs_vid.append(img.strip())
            else:
                imgs_pic.append(img.strip())
        if imgs_vid:
            try:
                post_id = post_facebook_video(group_id, imgs_vid,auth_token,title, status_message,
                    date, rating, address)
                imgs_id.append(post_id['id'])
            except AttributeError  as error:
                print("    An error occurred:",error)
                return False
        time.sleep(env.facebooksleep)
        print('    Facebook response: ',post_id)
    else:
        print ('    facebook: Wordpress URL or no pictures so skipping Facebook posting')
    return True

##################################################################################################

def post_to_threads2(title, content, headers, date, rating, address, picslist, local_outputs):
    """
    Posts content to threads if picslist is not empty and contains 'montage.mp4'.

    Args:
        title (str): The title of the post.
        content (str): The content of the post.
        headers: Headers information.
        date: Date information.
        rating: Rating information.
        address: Address information.
        picslist: List of pictures.
        local_outputs: Local output information.

    Returns:
        bool: True if the post was successfully uploaded to Instagram, False otherwise.
    """

    if picslist != '[]' and "montage.mp4" in picslist:
        outputmontage = ''
        addresshtml = re.sub(" ", ".",address)
        #content = content + get_hastags(address, title)
        pics = ((picslist[1:-1].replace(",","")).replace("'","")).split(" ")
        video, outputmontage = make_montage_video_from_google(pics)
        try:
            data =  title + "\n"+ address+"\nGoogle map to destination: " \
                r"https://www.google.com/maps/dir/?api=1&destination="+addresshtml +"\n\n" \
                + content + "\n"+rating+"\n"+date+"\n\n"+ get_hastags(address, title,'long')+ \
                "\n\nhttps://www.joeeatswhat.com "+"\n\n"
            local_outputs['instasession'].video_upload(outputmontage, data)
        except AttributeError  as error:
            print("  An error occurred uploading video to Threads:", type(error).__name__)
            return False
        return True
    else:
        return False

###################################################################################################

# def tiktok_upload_video(session_id, file_path, title, tags, schedule_time=None):
#     url = 'https://www.tiktok.com/api/upload/video/'
#     headers = {
#         'Cookie': f'sessionid={session_id}'
#     }
#     data = {
#         'title': title,
#         'tags': ','.join(tags),
#         'schedule_time': schedule_time
#     }
#     files = {
#         'video': open(file_path, 'rb')
#     }
#     response = requests.post(url, headers=headers, data=data, files=files)
#     return response.json()

def post_to_tiktok(title, content, headers, date, rating, address, picslist, local_outputs):
    """
    Posts content to TikTok with specified session ID, video file, title, hashtags, and optional
        scheduling.

    Args:
        title (str): The title of the video.
        content (str): The content of the video.
        headers: Headers information.
        date: Date information.
        rating: Rating information.
        address: Address information.
        picslist: List of pictures.
        local_outputs: Local output information.

    Returns:
        None
    """
    # Replace 'your_sessionid_cookie' with your actual TikTok sessionid cookie.
    #session_id = env.tiktok_client_secret

    # Replace 'path_to_video.mp4' with the path to your video file.
    #file_path = 'path_to_video.mp4'

    # Replace 'Your video title' with the title of your video.
    #title = 'Your video title'

    # Replace the following list with the hashtags you want to add to your post.
    tags = get_hastags(address, title,'long')

    # If you want to schedule your video, replace 'schedule_timestamp' with the Unix timestamp.
    # Leave it as None if you want to upload immediately.
    schedule_time = None  # or Unix timestamp (e.g., 1672592400)
    if picslist != '[]' and "montage.mp4" in picslist:
        for pic in picslist:
            if 'montage.mp4' in pic:
                file_path = pic
        #content = content + get_hastags(address, title)
        #pics = ((picslist[1:-1].replace(",","")).replace("'","")).split(" ")
#        video, outputmontage = make_montage_video_from_google(pics)
        # try:
        #     instasession.video_upload(outputmontage, data)
        # except AttributeError  as error:
        #     print("  An error occurred uploading video to Instagram:", type(error).__name__)
        #     return False
        # return True
"""
    client_key             string                        The unique identification key provisioned to the partner.
    client_secret         string                  The unique identification secret provisioned to the partner.
    code          string            The authorization code from the web, iOS, Android or desktop authorization callback. The value should be URL decoded.
    grant_type         string                 Its value should always be set as authorization_code.             
    redirect_uri    string               Its value must be the same as the redirect_uri used for requesting code.
    curl --location --request POST 'https://open.tiktokapis.com/v2/oauth/token/' \
    --header 'Content-Type: application/x-www-form-urlencoded' \
    --header 'Cache-Control: no-cache' \
    --data-urlencode 'client_key=CLIENT_KEY' \
    --data-urlencode 'client_secret=CLIENT_SECRET' \
    --data-urlencode 'code=CODE' \
    --data-urlencode 'grant_type=authorization_code' \
    --data-urlencode 'redirect_uri=REDIRECT_URI'
"""
    headers = ['Content-Type: application/x-www-form-urlencoded','Cache-Control: no-cache']
    data = ['client_key='+env.tiktok_client_key,'client_secret='+env.tiktok_client_secret,'code=CODE','grant_type=authorization_code','redirect_uri=REDIRECT_URI']
    response = requests.post('https://open.tiktokapis.com/v2/oauth/token/', headers=headers, data=data)

    # Call the function to upload the video
    # response = tiktok_upload_video(session_id, file_path, title, tags, schedule_time)
    print(response)
    return
###################################################################################################

def post_to_instagram2(title, content, headers, date, rating, address, picslist, local_outputs):
    """
    Posts content to Instagram with relevant information and media.

    Args:
        title (str): The title of the post.
        content (str): The content of the post.
        headers: Headers for the post.
        date: Date of the post.
        rating: Rating of the post.
        address: Address related to the post.
        picslist: List of pictures for the post.
        local_outputs: Local outputs for the post.

    Returns:
        bool: True if the post was successfully uploaded, False otherwise.
    """

    outputmontage = ''
    addresshtml = re.sub(" ", ".",address)
    attrib_list = local_outputs['postssession'].query(Posts).filter(Posts.name == title).all()
    if wpurl := attrib_list[0].wpurl:
        if business_url := attrib_list[0].businessurl:
            data =  title + "\n"+ address+"\n"+business_url+"\n"+"Review: "+wpurl+\
                "\nGoogle map to destination: " \
                r"https://www.google.com/maps/dir/?api=1&destination="\
                +addresshtml +"\n"+"Review: "+wpurl+"\n\n"+ content + "\n"+rating+"\n"+date+"\n\n"\
                +get_hastags(address, title,'long')+"\n\nhttps://www.joeeatswhat.com"+"\n\n"
        else:
            print ("    Missing business url for : "+title+" not using it in intagram post")
            data =  title + "\n"+ address+"\n"+"Review: "+wpurl+"\nGoogle map to destination: " \
                r"https://www.google.com/maps/dir/?api=1&destination="+addresshtml +"\nReview: "\
                +wpurl+"\n\n"+ content + "\n"+rating+"\n"+date+"\n\n"+ get_hastags(address, title,\
                'long')+"\n\nhttps://www.joeeatswhat.com"+"\n\n"
        instasession = local_outputs['instagram']
        if picslist != '[]' and "montage.mp4" in picslist:
            #content = content + get_hastags(address, title)
            pics = ((picslist[1:-1].replace(",","")).replace("'","")).split(" ")
            video, outputmontage = make_montage_video_from_google(pics)
            try:
                instasession.video_upload(outputmontage, data)
            except AttributeError  as error:
                print("  An error occurred uploading video to Instagram:", type(error).__name__)
                return False
            return True
    else:
        print ('    Missing wordpress post for instagram : ',title)
        return False

##################################################################################################

def post_to_wordpress(title,content,headers,date,rating,address,picslist,local_outputs):
    """
    Post to WordPress.

    This function posts content to a WordPress site using the provided data.

    Args:
        title (str): The title of the post.
        content (str): The content of the post.
        headers: Headers for the request.
        date (str): The date of the post.
        rating (str): The rating of the post.
        address (str): The address associated with the post.
        picslist (list): A list of pictures for the post.
        local_outputs: Outputs for the post.

    Returns:
        None
"""
    # post
    #new_post = False
    #countreview = False
    addresshtml = re.sub(" ", ".",address)
    googleadress = r"<a href=https://www.google.com/maps/dir/?api=1&destination="+\
            addresshtml + r">"+str(address)+r"</a>"
    contentpics = ''
    picchop = ''
    linkslist=[]
    # picl = picslist[1:-1]
    # pic2 = picl.replace(",","")
    # # #re.sub(r',','',picl) #re.sub( r'[^a-zA-Z0-9]','',tempdate[1])
    # pic3= pic2.replace("'","")
    # picchop = pic3.split(" ")
    # line = re.sub('[!@#$]', '', line)
    #picchop = re.sub(',\'','',picslist[1:-1])
    picchop = picslist[1:-1].replace(",","").replace("'","").split(" ")
    print ('    Figuring out date of Post : ',title)
    #specifify the formatting of the date_string.
    # formatting = '%b/%Y/%d'
    date_string = date
    if "a day" in date_string:
        date = dt.timedelta(days=-1)
#        newdate = dt.datetime.strptime(date_string, formatting).date()
        newdate = datetime.today() - date
        visitdate = newdate.strftime("%b%Y")
    else:
        if "day" in date:
            tempdate = -(int(re.sub( r'[^0-9]','',date_string)))
            print ('Stuff - > ',tempdate)
#           date = dt.timedelta(days=tempdate)
#            newdate = dt.datetime.strptime(date_string, formatting).date()
            newdate = datetime.today() + relativedelta(days=tempdate)
            visitdate = newdate.strftime("%b%Y")
        else:
            if "a week" in date:
#               date = dt.timedelta(weeks= -1)
#                newdate = dt.datetime.strptime(date_string, formatting).date()
                newdate = datetime.today() - relativedelta(weeks= 1)
                visitdate = newdate.strftime("%b%Y")
            else:
                if "week" in date:
                    tempdate = -(int(re.sub( r'[^0-9]','',date_string)))
                    print ('Stuff - > ',tempdate)
#                   date = dt.timedelta(weeks= tempdate)
#                    newdate = dt.datetime.strptime(date_string, formatting).date()
                    newdate = datetime.today() + relativedelta(weeks= tempdate)
                    visitdate = newdate.strftime("%b%Y")
                else:
                    if "a month" in date:
#                       date = dt.timedelta(months= -1)
#                       newdate = dt.datetime.strptime(date_string, formatting).date()
                        newdate = datetime.today() - relativedelta(months = 1)
                        visitdate = newdate.strftime("%b%Y")
                    else:
                        if "month" in date:
                            tempdate = -int(re.sub( r'[^0-9]','',date_string))
                            print ('Stuff - > ',tempdate)
#                           date = dt.timedelta(months= tempdate)
#                           newdate = dt.datetime.strptime(date_string, formatting).date()
                            newdate = datetime.today() + relativedelta(months =  tempdate)
                            visitdate = newdate.strftime("%b%Y")
                        else:
                            if "a year" in date:
#                               date = dt.timedelta(years= -1)
#                               newdate = dt.datetime.strptime(date_string, formatting).date()
                                newdate = datetime.today() - relativedelta(years= 1)
                                visitdate = newdate.strftime("%b%Y")
                            else:
                                if "year" in date:
                                    try:
                                        tempdate = -int(re.sub( r'[^0-9]','',date_string))
                                        print ('Stuff - > ',tempdate)
#                                       date = dt.timedelta( years= tempdate)
#                                       newdate = dt.datetime.strptime(date_string).date()
                                        newdate = datetime.today() + relativedelta(years= tempdate)
                                        visitdate = newdate.strftime("%b%Y")
                                    except AttributeError  as error:
                                        print("    An error getting date occurred:",error)
                                else:
                                    #specifify the formatting of the date_string.
                                    formatting = '%Y-%b-%d'
                                    month = date[:3]
                                    year = date[3:]
                                    day = '01'
                                    date_string = year+'-'+ month+'-'+day
                                    try:
                                        newdate = dt.datetime.strptime(date_string, formatting)\
                                                .date()
                                    except AttributeError  as error:
                                        print("    An error getting date occurred:",error)
#                                    try:
#                                        newdate = dt.datetime.strptime(date_string, formatting)\
#                                               .date()
#                                    except AttributeError as error:
#                                        print("    An error getting date occurred:", error)
                                    newdate = str(newdate)
    #formatting = '%b/%Y/%d' #specifify the formatting of the date_string.
    #newdate2 = dt.datetime.strptime(str(newdate), formatting).date()
    dateparts = (str(newdate)).split("-")
    dateparts2 = dateparts[2].split(" ")
    visitdate2 = local_outputs['postssession'].query(Posts).filter(Posts.name == title).all()[0].visitdate
    if visitdate != visitdate2:
        database_update_row(name,"visitdate",visitdate,"forceall",local_outputs)
        print (f'UPDATED: {visitdate2} to {visitdate} for {title}')
    #dateparts = dateparts2[0]
#    print ('dateparts',dateparts)
    newdate2 = dateparts[0]+'-'+dateparts[1]+'-'+dateparts2[0]+'T22:00:00'
    #newdate2 = str(re.sub(r'-','/',str(newdate.date())))+'T22:00:00'
    print ('    Got Date: ', newdate2, newdate)
    post_id, post_link = check_wordpress_post(title,visitdate,headers)
    featured_photo_id = get_wordpress_featured_photo_id(post_id)
    print(f"    Featured photo ID:  for {title} post {post_id} is: {featured_photo_id}")
    if env.block_google_maps is not True:
        try:
            database_update_row(title,"wpurl",post_link,"forceall",local_outputs)
        except  AttributeError  as error :
            print ('Could not check to see post already exists',error)
    if not post_id:
        googleadress =  r"<a href=https://www.google.com/maps/dir/?api=1&destination="+addresshtml\
            + r">"+address+r"</a>"
        post_data = {
            "title": title,
    #        "content": address+'\n\n'+content+'\n'+rating+'\n\n' ,
            "content": googleadress+'\n\n'+content+'\n'+rating ,
            "status": "publish",  # Set to 'draft' if you want to save as a draft
            "date": newdate2,
    #       "date": str(newdate)+'T22:00:00',
        # "author":"joesteele"
        }
        try:
            headers2 = headers
            response = requests.post(env.wpAPOurl, json = post_data, headers=headers2,timeout=env.request_timeout)
            if response.status_code != 201:
                print ('Error: ',response, response.text)
            else:
                #new_post = True
                post_id_json = response.json()
                post_id = post_id_json.get('id')
                print ('    New post is has post_id = ',post_id)
        except AttributeError  as error:
            print("An error occurred:", type(error).__name__) # An error occurred:
        #postneedsupdate = True
    else:
        print ('    Post already existed: Post ID : ',post_id)
    for pic in picchop:
        if pic != '':
            picslice2 = pic.split("/")[-1]
            picslice = picslice2.split(".")
            picname = picslice[0]
            caption =title
            description = title+"\n"+address
            print ('    Found Picture: ',picname)
            file_id, link = check_wordpress_media(picname, headers)
    #        link = linknew['rendered']
            if file_id is False:
                print ('      '+str(picname)+' was not already found in library, adding it')
    #            countreview = True
                image = {
                    "file": open(pic, "rb"),
                    "post": post_id,
                    "caption": caption,
                    "description": description
                }
                try:
                    image_response = requests.post(env.wpAPI + "/media", headers=headers, \
                        files=image,timeout=env.request_timeout)
                except AttributeError  as error:
                    print("    An error uploading picture ' + picname+ ' occurred:", \
                        type(error).__name__)
                if image_response.status_code != 201 :
                    print ('      Error:Image ',picname,' was not successfully uploaded.  response: ',\
                        image_response)
                else:
                    pic_dic=image_response.json()
                    file_id= pic_dic.get('id')
                    link = pic_dic.get('guid').get("rendered")
                    print ('      ',picname,' was successfully uploaded to website with ID: ',\
                        file_id, link)
                try:
                    links_dict = {'file_id' : file_id , 'link' : link}
                    linkslist.append(links_dict)
                except AttributeError  as error:
                    print("    An error adding to dictionary " , file_id , link , " occurred:",\
                        type(error).__name__) # An error occurred:
            else:
                print ('    Photo ',picname,' was already in library and added to post with ID: ',\
                    file_id,' : ',link)
                try:
                    image_response = requests.post(env.wpAPI + "/media/" + str(file_id),\
                        headers=headers, data={"post" : post_id},timeout=env.request_timeout)
                except AttributeError  as error:
                    print ('    Error- Image ',picname,' was not attached to post.  response: ',\
                        image_response+' '+type(error).__name__)
                try:
                    post_response = requests.post(env.wpAPI + "/posts/" + str(post_id),\
                        headers=headers,timeout=env.request_timeout)
                    if link in post_response.text:
                        print ('    Image link for ', picname, 'already in content of post: '\
                            ,post_id, post_response.text, link)
                    else:
                        linkslist.append({'file_id' : file_id , 'link' : link})
    #                   countreview = True
                except AttributeError  as error:
                    print("    An error loading the metadata from the post "+post_response.title+\
                        ' occurred: '+type(error).__name__)
    #ratinghtml = post_response.text
    first_mp4 = True
    first_pic = True
    fmedia = {}
    contentpics = ""
    for piclink in linkslist:
        if piclink != '':
            #for loop in linkslist:
            print ('    Adding ', piclink['link'], ' to posting')
            try:
                ext = piclink['link'].split( '.')[-1]
                if ext == 'mp4':
                    if first_mp4:
                        contentpics += '\n' +r'[evp_embed_video url="' + piclink['link'] + \
                            r'" autoplay="true"]'
                        first_mp4 = False
                    else:
                        contentpics += '\n' +r'[evp_embed_video url="' + piclink['link'] + r'"]'
    #[evp_embed_video url="http://example.com/wp-content/uploads/videos/vid1.mp4" autoplay="true"]
                else:
                    contentpics+='\n '+r'<div class="col-xs-4"><img id="'+str(file_id)+r'"'+r'src="'+\
                        piclink['link'] + r'"></div>'
                    if first_pic:
                        fmedia = piclink['file_id']
                        first_pic = False
    #               fmedia.append = piclink{'file_id' }
    #            contentpics += '\n '+r'<img src="'+ piclink['link'] + '> \n'
                #contentpics += r'<img src="'+ piclink['link'] + r' alt="' + title +r'">' +'\n\n'
            except AttributeError  as error:
                print("An error occurred:", type(error).__name__) # An error occurred:
                return False
    try:
#        print ('featured_media = ',linkslist[0]['file_id'])
        # if linkslist[0]['file_id']:
        #     print ('featuredmedia2 = ',linkslist[0]['file_id'])
        if fmedia:
            print ('Featured Media: ',fmedia)
        else:
            if linkslist:
                fmedia = linkslist[0]['file_id']
#            print ('featured_media2 = ',file_id)
        business_url_list = local_outputs['postssession'].query(Posts).filter(Posts.name == title)\
                .all()
        business_url = "<a href="+str(business_url_list[0].businessurl)+">"+str(business_url_list[0].businessurl)+"</a>"
        # wpurllist = local_outputs['postssession'].query(Posts).filter(Posts.name == title).all()
        # wpurl = wpurllist[0].wpurl
#        if business_url or business_url is False:
        status_message = str(title) + ': Business website: '+ business_url
        response_piclinks = requests.post(env.wpAPI+"/posts/"+ str(post_id), \
            data={"content" : title+' - '+status_message+'\n\n'+content+'\n'+googleadress+'\n'+\
            rating+contentpics,"featured_media":fmedia,"rank_math_focus_keyword":title},\
            headers=headers,timeout=env.request_timeout)
        print ('  ',response_piclinks)
        if fmedia and fmedia != 0 and featured_photo_id and featured_photo_id !=0:
            print ("    Featured Media: "+str(featured_photo_id)+" "+str(fmedia)+"  looks OK")
        else:
            print ("    fmedia is still empty and need to populate")
            if fmedia:
                response_piclinks = requests.post(env.wpAPI+"/posts/"+ str(post_id), \
                    data={"content" : title+' - '+status_message+'\n\n'+content+'\n'+googleadress+'\n'+\
                    rating+contentpics,"featured_media":fmedia,"rank_math_focus_keyword":title},\
                    headers=headers,timeout=env.request_timeout)
                print ('    Results of new results ID value : '+str(get_wordpress_featured_photo_id(post_id)))
            else:
                response_piclinks = requests.post(env.wpAPI+"/posts/"+ str(post_id), \
                    data={"content" : title+' - '+status_message+'\n\n'+content+'\n'+googleadress+'\n'+\
                    rating+contentpics,"rank_math_focus_keyword":title},\
                    headers=headers,timeout=env.request_timeout)
                print ('Results of new results ID  with no picture value : '+str(get_wordpress_featured_photo_id(post_id)))
    except AttributeError  as error:
        print("    An error writing images to the post " + post_response.title + ' occurred:',\
            type(error).__name__) # An error occurred')
        return False
    return True

##################################################################################################

def build_picslist(picchop, piclink):
    """
    Builds a list of pictures for posting based on the provided picture links.

    Args:
        picchop (list): List of picture links to process.
        piclink (str): The picture link to be added to the list.

    Returns:
        str: A string containing the formatted list of pictures for posting.
    Raises:
        AttributeError: If an error occurs during picture processing.
    """

    for piclink in picchop:
        if piclink != '':
            print ('    Adding ', piclink['link'], ' to posting')
            try:
                ext = piclink['link'].split( '.')[-1]
                if ext == 'mp4':
                    if first_mp4:
                        contentpics += '\n' +r'[evp_embed_video url="' + piclink['link'] + \
                            r'" autoplay="true"]'
                        first_mp4 = False
                    else:
                        contentpics += '\n' +r'[evp_embed_video url="' + piclink['link'] + r'"]'
                else:
                    contentpics+='\n '+r'<div class="col-xs-4"><img id="'+str(file_id)+r'"'+r'src="'+\
                        piclink['link'] + r'"></div>'
                    if first_pic:
                        fmedia = piclink['file_id']
                        first_pic = False
            except AttributeError  as error:
                print("An error occurred:", type(error).__name__) # An error occurred:
                return False
    return contentpics

##################################################################################################

def get_wordpress_post_date_string(date_string, date):
    """
    Parses a date string and returns two formatted date strings.

    Args:
        date_string (str): The input date string to be parsed.
        date (str): The keyword indicating the date offset.

    Returns:
        tuple: A tuple containing two formatted date strings.
    Raises:
        AttributeError: If an error occurs during date parsing.
    """


    if "a day" in date_string:
        date = dt.timedelta(days=-1)
        newdate = datetime.today() - date
        visitdate = newdate.strftime("%b%Y")
    else:
        if "day" in date:
            tempdate = -(int(re.sub( r'[^0-9]','',date_string)))
            print ('Stuff - > ',tempdate)
            newdate = datetime.today() + relativedelta(days=tempdate)
            visitdate = newdate.strftime("%b%Y")
        else:
            if "a week" in date:
                newdate = datetime.today() - relativedelta(weeks= 1)
                visitdate = newdate.strftime("%b%Y")
            else:
                if "week" in date:
                    tempdate = -(int(re.sub( r'[^0-9]','',date_string)))
                    print ('Stuff - > ',tempdate)
                    newdate = datetime.today() + relativedelta(weeks= tempdate)
                    visitdate = newdate.strftime("%b%Y")
                else:
                    if "a month" in date:
                        newdate = datetime.today() - relativedelta(months = 1)
                        visitdate = newdate.strftime("%b%Y")
                    else:
                        if "month" in date:
                            tempdate = -int(re.sub( r'[^0-9]','',date_string))
                            print ('Stuff - > ',tempdate)
                            newdate = datetime.today() + relativedelta(months =  tempdate)
                            visitdate = newdate.strftime("%b%Y")
                        else:
                            if "a year" in date:
                                newdate = datetime.today() - relativedelta(years= 1)
                                visitdate = newdate.strftime("%b%Y")
                            else:
                                if "year" in date:
                                    try:
                                        tempdate = -int(re.sub( r'[^0-9]','',date_string))
                                        print ('Stuff - > ',tempdate)
                                        newdate = datetime.today() + relativedelta(years= tempdate)
                                        visitdate = newdate.strftime("%b%Y")
                                    except AttributeError  as error:
                                        print("    An error getting date occurred:",error)
                                else:
                                    #specifify the formatting of the date_string.
                                    formatting = '%Y-%b-%d'
                                    month = date[:3]
                                    year = date[3:]
                                    day = '01'
                                    date_string = year+'-'+ month+'-'+day
                                    try:
                                        newdate = dt.datetime.strptime(date_string, formatting)\
                                                .date()
                                    except AttributeError  as error:
                                        print("    An error getting date occurred:",error)
                                    newdate = str(newdate)
    #formatting = '%b/%Y/%d' #specifify the formatting of the date_string.
    #newdate2 = dt.datetime.strptime(str(newdate), formatting).date()
    dateparts = (str(newdate)).split("-")
    dateparts2 = dateparts[2].split(" ")
    newdate2 = dateparts[0]+'-'+dateparts[1]+'-'+dateparts2[0]+'T22:00:00'
    print ('    Got Date: ', newdate2, newdate)
    return (newdate,newdate2,visitdate)

##################################################################################################

def create_wordpress_post(newdate2, picchop, title, address, headers, post_id, linkslist, local_outputs, content, googleadress, rating, featured_photo_id):
    """
    Creates a WordPress post with pictures and content.

    Args:
        newdate2: The date of the post.
        picchop (list): List of picture links to process.
        title: The title of the post.
        address: The address associated with the post.
        headers: The headers for the HTTP request.
        post_id: The ID of the post.
        linkslist: List of links to pictures.
        local_outputs: Local outputs for the post.
        content: The content of the post.
        googleadress: The Google address for the post.
        rating: The rating of the post.
        featured_photo_id: The ID of the featured photo.

    Returns:
        bool: True if the post creation is successful, False otherwise.
    Raises:
        AttributeError: If an error occurs during post creation or processing.
    """

    if not post_id:
        googleadress =  r"<a href=https://www.google.com/maps/dir/?api=1&destination="+addresshtml\
            + r">"+address+r"</a>"
        post_data = {
            "title": title,
            "content": googleadress+'\n\n'+content+'\n'+rating ,
            "status": "publish",  # Set to 'draft' if you want to save as a draft
            "date": newdate2
        }
        try:
            headers2 = headers
            response = requests.post(env.wpAPOurl, json = post_data, headers=headers2,timeout=env.request_timeout)
            if response.status_code != 201:
                print ('Error: ',response, response.text)
            else:
                post_id_json = response.json()
                post_id = post_id_json.get('id')
                print ('    New post has post_id = ',post_id)
        except AttributeError  as error:
            print("An error occurred:", type(error).__name__) # An error occurred:
    else:
        print ('    Post already existed: Post ID : ',post_id)
    for pic in picchop:
        if pic != '':
            picslice2 = pic.split("/")[-1]
            picslice = picslice2.split(".")
            picname = picslice[0]
            caption =title
            description = title+"\n"+address
            print ('    Found Picture: ',picname)
            file_id, link = check_wordpress_media(picname, headers)
            if file_id is False:
                print(f'      {str(picname)} was not already found in library, adding it')
                image = {
                    "file": open(pic, "rb"),
                    "post": post_id,
                    "caption": caption,
                    "description": description
                }
                try:
                    image_response = requests.post(env.wpAPI + "/media", headers=headers, \
                        files=image,timeout=env.request_timeout)
                except AttributeError  as error:
                    print(f"    An error uploading picture {picname} occurred:", \
                        type(error).__name__)
                if image_response.status_code != 201 :
                    print (f'      Error:Image {picname} was not successfully uploaded.  response: ',\
                        image_response)
                else:
                    pic_dic=image_response.json()
                    file_id= pic_dic.get('id')
                    link = pic_dic.get('guid').get("rendered")
                    print (f'      {picname} was successfully uploaded to website with ID: ',\
                        file_id, link)
                try:
                    links_dict = {'file_id' : file_id , 'link' : link}
                    linkslist.append(links_dict)
                except AttributeError  as error:
                    print(f"    An error adding to dictionary {file_id} {link} occurred:",\
                        type(error).__name__) # An error occurred:
            else:
                print (f'    Photo {picname} was already in library and added to post with ID: \
                    {file_id} : {link}')
                try:
                    image_response = requests.post(env.wpAPI + "/media/" + str(file_id),\
                        headers=headers, data={"post" : post_id},timeout=env.request_timeout)
                except AttributeError  as error:
                    print ('    Error- Image ',picname,' was not attached to post.  response: ',\
                        image_response+' '+type(error).__name__)
                try:
                    post_response = requests.post(env.wpAPI + "/posts/" + str(post_id),\
                        headers=headers,timeout=env.request_timeout)
                    if link in post_response.text:
                        print ('    Image link for ', picname, 'already in content of post: '\
                            ,post_id, post_response.text, link)
                    else:
                        linkslist.append({'file_id' : file_id , 'link' : link})
                except AttributeError  as error:
                    print("    An error loading the metadata from the post "+post_response.title+\
                        ' occurred: '+type(error).__name__)
    first_mp4 = True
    first_pic = True
    fmedia = {}
    contentpics = ""
    for piclink in linkslist:
        if piclink != '':
            print ('    Adding ', piclink['link'], ' to posting')
            try:
                ext = piclink['link'].split( '.')[-1]
                if ext == 'mp4':
                    if first_mp4:
                        contentpics += '\n' +r'[evp_embed_video url="' + piclink['link'] + \
                            r'" autoplay="true"]'
                        first_mp4 = False
                    else:
                        contentpics += '\n' +r'[evp_embed_video url="' + piclink['link'] + r'"]'
                else:
                    contentpics+='\n '+r'<div class="col-xs-4"><img id="'+str(file_id)+r'"'+r'src="'+\
                        piclink['link'] + r'"></div>'
                    if first_pic:
                        fmedia = piclink['file_id']
                        first_pic = False
            except AttributeError  as error:
                print("An error occurred:", type(error).__name__) # An error occurred:
                return False
    try:
        if fmedia:
            print ('Featured Media: ',fmedia)
        else:
            if linkslist:
                fmedia = linkslist[0]['file_id']
        business_url_list = local_outputs['postssession'].query(Posts).filter(Posts.name == title)\
                .all()
        business_url = f"<a href={str(business_url_list[0].businessurl)}>{title}</a>"
        status_message = f'{str(title)}: Business website: {business_url}'
        response_piclinks = requests.post(env.wpAPI+"/posts/"+ str(post_id), \
            data={"content" : title+' - '+status_message+'\n\n'+content+'\n'+googleadress+'\n'+\
            rating+contentpics,"featured_media":fmedia,"rank_math_focus_keyword":title},\
            headers=headers,timeout=env.request_timeout)
        print ('  ',response_piclinks)
        if fmedia and fmedia != 0 and featured_photo_id and featured_photo_id !=0:
            print ("    Featured Media: "+str(featured_photo_id)+" "+str(fmedia)+"  looks OK")
        else:
            print ("    fmedia is still empty and need to populate")
            if fmedia:
                response_piclinks = requests.post(env.wpAPI+"/posts/"+ str(post_id), \
                    data={"content" : title+' - '+status_message+'\n\n'+content+'\n'+googleadress+'\n'+\
                    rating+contentpics,"featured_media":fmedia,"rank_math_focus_keyword":title},\
                    headers=headers,timeout=env.request_timeout)
                print ('    Results of new results ID value : '+str(get_wordpress_featured_photo_id(post_id)))
            else:
                response_piclinks = requests.post(env.wpAPI+"/posts/"+ str(post_id), \
                    data={"content" : title+' - '+status_message+'\n\n'+content+'\n'+googleadress+'\n'+\
                    rating+contentpics,"rank_math_focus_keyword":title},\
                    headers=headers,timeout=env.request_timeout)
                print ('Results of new results ID  with no picture value : '+str(get_wordpress_featured_photo_id(post_id)))
    except AttributeError  as error:
        print("    An error writing images to the post " + post_response.title + ' occurred:',\
            type(error).__name__) # An error occurred')
        return False
    return True

##################################################################################################

def post_to_wordpress2(title,content,headers,date,rating,address,picslist,local_outputs):
    """
    Post to WordPress.

    This function posts content to a WordPress site using the provided data.

    Args:
        title (str): The title of the post.
        content (str): The content of the post.
        headers: Headers for the request.
        date (str): The date of the post.
        rating (str): The rating of the post.
        address (str): The address associated with the post.
        picslist (list): A list of pictures for the post.
        local_outputs: Outputs for the post.

    Returns:
        None
"""
    linkslist = ()
    addresshtml = re.sub(" ", ".",address)
    googleadress = r"<a href=https://www.google.com/maps/dir/?api=1&destination="+\
            addresshtml + r">"+str(address)+r"</a>"
    picchop = picslist[1:-1].replace(",","").replace("'","").split(" ")
    print ('    Figuring out date of Post : ',title)
    #specifify the formatting of the date_string.
    # formatting = '%b/%Y/%d'
    date_string = date
    newdate,newdate2,visitdate = get_wordpress_post_date_string(date_string,date)
    if picslist and picslist != '':
        content_pics = build_picslist(picchop,piclink)
        featured_photo_id = get_wordpress_featured_photo_id(post_id)
        print(f"    Featured photo ID:  for {title} post {post_id} is: {featured_photo_id}")
    if env.block_google_maps is not True or env.forcegoogleupdate is True:
        try:
            database_update_row(title,"wpurl",post_link,"forceall",local_outputs)
        except  AttributeError  as error :
            print (f'     Error: {error}')
    else:
        post_id, post_link = check_wordpress_post(title,visitdate,headers)
        if post_id:
            print ('    Post already existed: Post ID : ',post_id)
            print ('    Found post for : '+title)
            if env.force_web_create:
                create_wordpress_post(newdate2,picchop,title,address,headers,post_id,linkslist,local_outputs,content,googleadress,rating,featured_photo_id)
                #update_wordpress()
            else:
                print ('    Found existing post but skipping updating post')
        else:
            create_wordpress_post(newdate2,picchop,title,address,headers,post_id,linkslist,local_outputs,content,googleadress,rating,featured_photo_id)
            print ('    Creating wordpress post from scratch for: '+title)
    return True

##################################################################################################

def process_reviews2(outputs):
    """
    Processes reviews data, updates social media platforms, and writes data to files.

    Args:
        outputs: Data outputs containing reviews and social media information.

    Returns:
        None
    """

    # Process
    webcount = xtwittercount = instagramcount = facebookcount = 0
#    webcount=xtwittercount=instagramcount=yelpcount=threadscount=facebookcount=tiktokcount = 0
    if env.datasource == 'db':
        cols2 = ["num","name", "comment", 'rating','picsURL','picsLocalpath','source','date',
        'address','dictPostComplete']
        rows_orig = list(outputs['data'].iter_rows(min_row=1, max_row=outputs['data'].max_row))
        ttt = list(outputs['data'].iter_rows(min_row=1, max_row=outputs['data'].max_row))
        ows_orig = list(outputs['data'])
        rows_orig2 = list(outputs['xlsdf'].values)
        rows3 = outputs['data'].iter_rows(min_row=1, max_row=outputs['data'])
        rows4 = pd.DataFrame(outputs['xlsdf'])
        rows5 = pd.DataFrame(outputs['posts'])
        rows6 = pd.DataFrame(outputs['data'])
# df2 = pd.DataFrame(outputs['xlsdf'].values, columns=cols2).iter_rows(min_row=1, max_row=outputs['data'].max_row)
#   df3 = pd.DataFrame(outputs['xlsdf'].iter_rows(min_row=1, max_row=outputs['data'].max_row))
# rows = [({0:p.id},{1:p.name}, { 2:p.comment}, {3: p.rating}, {4:p.picsURL},\
#   {5:p.picsLocalpath},{6:p.source},{7:p.date},{8:p.address},{9:p.dictPostComplete})\
#   for p in rows_orig]
        # rows = list((outputs['data'].iter_rows(min_row=1, max_row=outputs['data'].max_row)))
        rows = outputs['posts']
    else:
        rows_orig = outputs['posts']
        rows = rows_orig
    if env.google:
        print('Configuration says to update google Reviews prior to processing them')
        options = webdriver.ChromeOptions()
        options.add_argument("--log-level=3")
        options.add_argument("--ignore-certificate-error")
        options.add_argument("--ignore-ssl-errors")
        if not env.showchrome:
            options.add_argument("--headless")
            # show browser or not ||| HEAD =>  43.03 ||| No Head => 39 seg
        options.add_argument("--lang=en-US")
        options.add_argument("--disable-blink-features=AutomationControlled")
        options.add_argument("--remote-debugging-pipe")
        # Exclude the collection of enable-automation switches
        options.add_experimental_option("excludeSwitches", ["enable-automation"])
        # Turn-off userAutomationExtension
        options.add_experimental_option("useAutomationExtension", False)
        # Setting the driver path and requesting a page
        caps = webdriver.DesiredCapabilities.CHROME.copy()
        caps['acceptInsecureCerts'] = True
        caps['acceptSslCerts'] = True
        options.set_capability('cloud:options', caps)
        #driver = webdriver.Chrome(desired_capabilities=caps)
        if is_docker():
            driver = webdriver.Remote("http://192.168.10.9:4444/wd/hub", options=options)
            print ("IN A DOCKER CONTAINER, USING REMOTE CHROME")
        else:
            driver = webdriver.Chrome(options=options) # Firefox(options=options)
        # Changing the property of the navigator value for webdriver to undefined
        driver.execute_script("Object.defineProperty(navigator,'webdriver',\
                            {get:()=> undefined})")
        driver.get(env.URL)
        time.sleep(5)
        google_scroll(counter_google(driver), driver)
        webdata = get_google_data(driver,outputs)
        if env.datasource == 'db':
            write_to_database(webdata, outputs)
        else:
            write_to_xlsx2(webdata, outputs)
        driver.close()
        # outputs['data'].save(xls)
        print('Done getting google reviews and writing them to xls file !')
    else:
        print ('Configuration says to skip creation of new reviews from google for this run')
    print('Processing Reviews')
    for processrow in rows:
        if processrow.name != "name":  # Skip header line of xls sheet
            outputs['postssession'].query(Posts).filter(Posts.name == processrow.name).\
                    update({"google" : 1})
            print ("Processing : ",processrow.name)
            writtento = ast.literal_eval(processrow.dictPostComplete)
            # Check to see if the website has already been written to according to the xls sheet,\
            # if it has not... then process
            if (writtento["web"] == 0 or writtento["instagram"]==0 or writtento["facebook"]==0 or \
                writtento["xtwitter"]==0 or writtento["yelp"]==0 or writtento["tiktok"]==0 or \
                writtento["threads"]==0 ) and (check_is_port_open(env.wpAPI, 443)) and (env.web \
                or env.instagram or env.yelp or env.xtwitter or env.tiktok or env.facebook or \
                env.threads or env.google)and (processrow.comment is not None) :
                if env.web  and processrow.web is False or env.force_web_create is True:
                    #if writtento["web"] == 0 :
                    try:
                        post_id, post_link = get_wordpress_post_id_and_link(processrow.name,\
                                outputs['web'] )
                        if env.forcegoogleupdate is True and env.block_google_maps is not True:
                            if post_link:
                                database_update_row(processrow.name,"wpurl",post_link,"forceall"\
                                        ,outputs)
                            else:
                                print ('  Error getting wordpress links to update databse')
                    except  AttributeError  as error :
                        print ('Could not check to see post already exists',error)
                    webcount=process_socials("web",processrow,outputs['web'],"post_to_wordpress",\
                            webcount, outputs)
                if env.instagram and processrow.instagram is False:
                    instagramcount = process_socials("instagram",processrow,outputs['web'],\
                            "post_to_instagram2",instagramcount, outputs)
                if env.facebook and processrow.facebook is False  :
                    facebookcount = process_socials("facebook",processrow,outputs['web'],\
                            "post_facebook3",facebookcount, outputs)            
                if env.xtwitter and processrow.xtwitter is False:
                    xtwittercount = process_socials("xtwitter",processrow,outputs['web'],\
                            "post_to_x2",xtwittercount, outputs)
    return

##################################################################################################

def process_socials(social_name,social_post,headers,sub_process,social_count, local_outputs):
    """Summary:
    Function to process social media posts.

    Explanation:
    This function processes social media posts based on the provided parameters.

    Args:
    - social_name: The name of the social media platform.
    - social_post: The post content for the social media platform.
    - sub_process: The sub-process for posting.
    - social_count: The count of social media posts.
    - local_outputs: Dictionary containing outputs from the process.

    Returns:
    Count of the social that was selected
    """
    writtento = ast.literal_eval(social_post.dictPostComplete)
    if (len(local_outputs['postssession'].query(Posts).filter(Posts.name == social_post.name,\
            getattr(Posts, social_name) is True).all())==0) and ((local_outputs['postssession'].\
                    query(Posts).filter(Posts.name == social_post.name).all()[0].wpurl is not \
                    None)or social_name == 'web'):
        if social_count < env.postsperrun or (social_name == 'web' and env.force_web_create is True):
            try:
                print('  Starting to generate ',social_name,' post')
                new_social_post = eval(sub_process)(social_post.name, social_post.comment,\
                        headers, social_post.date, social_post.rating, social_post.address,\
                        social_post.picsLocalpath,local_outputs )
                try:
                    print ('    Start generating content to post to : ',social_post.name)
                    writtento[social_name] = 1
                    social_post.dictPostComplete = str(writtento)
                except AttributeError  as error:
                    print("  An error occurred setting value to go into Excel file:", type(error)\
                            .__name__)
                    print ('  Success Posting to xtwitter: '+social_post.name)
                if new_social_post:
                    social_count +=1
                    try:
                        print('  write to xls for :',social_name)
                        local_outputs['datawb'].save(env.xls)
                        print('  Successfully wrote to xls for social - ',social_name)
                    except AttributeError  as error:
                        print("  An error occurred writing Excel file:", type(error).__name__)
                    try:
                        print('  write to DB for Social - ',social_name)
                        local_outputs['postssession'].query(Posts).filter(Posts.name == \
                            social_post.name).update({social_name : True})
                        local_outputs['postssession'].commit()
                        print('  Successfully wrote to database')
                    except AttributeError  as error:
                        print("  An error occurred writing database", type(error).__name__)
            except AttributeError as error:
                print('  Error writing social - ',social_name,'  post : ',error,social_post.name,\
                    social_post.comment, local_outputs,social_post.date, social_post.rating,\
                    social_post.address, social_post.picsLocalpath, writtento[social_name],\
                    type(error).__name__ )
        else:
            print ('  Exceeded the number of social - ',social_name,' posts per run, skipping',\
                    social_post.name)
    else:
        print ('  ',social_name,': Skipping posting for ',social_post.name,' previously written')
    return social_count

##################################################################################################

if __name__ == "__main__":
    print('starting ...')
    preload()
    print('making connections ...')
    outputs = get_auth_connect()
    process_reviews2(outputs)
    print('Done!')

##################################################################################################
